import calendar
import locale
import os
from datetime import datetime, date
from urllib.parse import quote

import openpyxl
from django.db.models import Q
from django.http import HttpResponse
from django.utils import translation
from django.utils.formats import date_format
from openpyxl import load_workbook

from apps.core.models import Employee, Branch
from .models import VisitReport, BookReport, Event, VisitsFirstData, EventsFirstData, BooksFirstData, VisitPlan, \
    EventsPlan, RegsPlan, \
    BookPlan


def generate_all_reports_excel(user, year, month):
    # Получаем сотрудника и филиал
    try:
        employee = Employee.objects.get(user=user)
        branch = employee.branch
    except Employee.DoesNotExist:
        return None

    if not branch:
        return None

    # Загружаем общий шаблон
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(base_dir, 'reports/excell/report_all_template.xlsx')

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"File not found: {template_path}")

    wb = openpyxl.load_workbook(filename=template_path)

    # Генерация отчёта по посещениям
    generate_visit_report(wb, branch, year, month)

    # Генерация отчёта по книговыдаче
    generate_book_report(wb, branch, year, month)

    # Генерация отчёта по мероприятиям
    generate_events_report(wb, branch, year, month)

    # Сохраняем файл в HttpResponse
    month_name = calendar.month_name[month]
    filename = f"эл.дневник_{branch.short_name}_{year}_{month_name}.xlsx"
    safe_filename = quote(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
    wb.save(response)
    return response


def generate_visit_report(wb, branch, year, month):
    ws = wb["Число пользователей+посещения"]

    # Фильтруем отчеты по филиалу, году и месяцу
    visit_reports = VisitReport.objects.filter(
        library=branch,
        date__year=year,
        date__month=month
    ).order_by('date')

    # Фильтруем отчеты с начала года до текущего месяца (не включая текущий месяц)
    visit_reports_year = VisitReport.objects.filter(
        library=branch,
        date__year=year,
        date__month__lt=month
    ).order_by('date')

    # Агрегируем данные с начала года
    year_start_data = {
        'qty_reg_14': sum(report.qty_reg_14 for report in visit_reports_year),
        'qty_reg_15_35': sum(report.qty_reg_15_35 for report in visit_reports_year),
        'qty_reg_18_35': sum(report.qty_reg_18_35 for report in visit_reports_year),
        'qty_reg_other': sum(report.qty_reg_other for report in visit_reports_year),
        'qty_reg_pensioners': sum(report.qty_reg_pensioners for report in visit_reports_year),
        'qty_reg_invalid': sum(report.qty_reg_invalid for report in visit_reports_year),
        'qty_reg_out_of_station': sum(report.qty_reg_out_of_station for report in visit_reports_year),
        'qty_reg_prlib': sum(report.qty_reg_prlib for report in visit_reports_year),
        'qty_reg_litres': sum(report.qty_reg_litres for report in visit_reports_year),
        'qty_visited_14': sum(report.qty_visited_14 for report in visit_reports_year),
        'qty_visited_15_35': sum(report.qty_visited_15_35 for report in visit_reports_year),
        'qty_visited_18_35': sum(report.qty_visited_18_35 for report in visit_reports_year),
        'qty_visited_other': sum(report.qty_visited_other for report in visit_reports_year),
        'qty_visited_pensioners': sum(report.qty_visited_pensioners for report in visit_reports_year),
        'qty_visited_invalids': sum(report.qty_visited_invalids for report in visit_reports_year),
        'qty_visited_out_station': sum(report.qty_visited_out_station for report in visit_reports_year),
        'qty_visited_prlib': sum(report.qty_visited_prlib for report in visit_reports_year),
        'qty_visited_litres': sum(report.qty_visited_litres for report in visit_reports_year),
        'qty_visited_online': sum(report.qty_visited_online for report in visit_reports_year),
    }

    # Записываем данные с начала года в строку 6
    ws['B6'] = (year_start_data['qty_reg_14'] + year_start_data['qty_reg_15_35'] + year_start_data['qty_reg_18_35']
                + year_start_data['qty_reg_other'])
    ws['C6'] = year_start_data['qty_reg_14']
    ws['D6'] = year_start_data['qty_reg_15_35']
    ws['E6'] = year_start_data['qty_reg_18_35']
    ws['F6'] = year_start_data['qty_reg_other']
    ws['G6'] = year_start_data['qty_reg_pensioners']
    ws['H6'] = year_start_data['qty_reg_invalid']
    ws['I6'] = year_start_data['qty_reg_out_of_station']
    ws['J6'] = year_start_data['qty_reg_prlib']
    ws['K6'] = year_start_data['qty_reg_litres']
    ws['L6'] = (year_start_data['qty_visited_14'] + year_start_data['qty_visited_15_35'] + year_start_data['qty_visited_18_35']
                + year_start_data['qty_visited_other'])
    ws['M6'] = year_start_data['qty_visited_14']
    ws['N6'] = year_start_data['qty_visited_15_35']
    ws['O6'] = year_start_data['qty_visited_18_35']
    ws['P6'] = year_start_data['qty_visited_other']
    ws['Q6'] = year_start_data['qty_visited_pensioners']
    ws['R6'] = year_start_data['qty_visited_invalids']
    ws['S6'] = year_start_data['qty_visited_out_station']
    ws['T6'] = year_start_data['qty_visited_prlib']
    ws['U6'] = year_start_data['qty_visited_litres']
    ws['V6'] = year_start_data['qty_visited_online']

    # Заполняем данные за текущий месяц
    current_date = date(year, month, 1)
    month_name_ru = date_format(current_date, "F")
    ws['A1'] = f"Отчет по посещениям за {month_name_ru} {year} года"
    ws['O1'] = f"{branch}"

    row_num = 7  # Начинаем с 7 строки (после заголовков)
    for report in visit_reports:
        # Вставляем строку перед записью данных (кроме первой строки)
        if row_num > 7:
            ws.insert_rows(row_num)  # Вставляем новую строку
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=col)._style = ws.cell(row=7, column=col)._style  # Копируем стили

        # Записываем данные в текущую строку
        ws[f'A{row_num}'] = report.date.strftime('%d.%m.%Y')
        ws[f'B{row_num}'] = report.qty_reg_14 + report.qty_reg_15_35 + report.qty_reg_18_35 + report.qty_reg_other
        ws[f'C{row_num}'] = report.qty_reg_14
        ws[f'D{row_num}'] = report.qty_reg_15_35
        ws[f'E{row_num}'] = report.qty_reg_18_35
        ws[f'F{row_num}'] = report.qty_reg_other
        ws[f'G{row_num}'] = report.qty_reg_pensioners
        ws[f'H{row_num}'] = report.qty_reg_invalid
        ws[f'I{row_num}'] = report.qty_reg_out_of_station
        ws[f'J{row_num}'] = report.qty_reg_prlib
        ws[f'K{row_num}'] = report.qty_reg_litres
        ws[f'L{row_num}'] = report.qty_visited_14 + report.qty_visited_15_35 + report.qty_visited_18_35 + report.qty_visited_other
        ws[f'M{row_num}'] = report.qty_visited_14
        ws[f'N{row_num}'] = report.qty_visited_15_35
        ws[f'O{row_num}'] = report.qty_visited_18_35
        ws[f'P{row_num}'] = report.qty_visited_other
        ws[f'Q{row_num}'] = report.qty_visited_pensioners
        ws[f'R{row_num}'] = report.qty_visited_invalids
        ws[f'S{row_num}'] = report.qty_visited_out_station
        ws[f'T{row_num}'] = report.qty_visited_prlib
        ws[f'U{row_num}'] = report.qty_visited_litres
        ws[f'V{row_num}'] = report.qty_visited_online
        ws[f'W{row_num}'] = report.note
        row_num += 1


def generate_book_report(wb, branch, year, month):
    ws = wb["Книговыдача"]

    # Фильтруем отчеты по филиалу, году и месяцу
    book_reports = BookReport.objects.filter(
        library=branch,
        date__year=year,
        date__month=month
    ).order_by('date')

    # Фильтруем отчеты с начала года до текущего месяца (не включая текущий месяц)
    book_reports_year = BookReport.objects.filter(
        library=branch,
        date__year=year,
        date__month__lt=month
    ).order_by('date')

    # Агрегируем данные с начала года
    year_start_data = {
        'qty_books_14': sum(report.qty_books_14 for report in book_reports_year),
        'qty_books_15_35': sum(report.qty_books_15_35 for report in book_reports_year),
        'qty_books_18_35': sum(report.qty_books_18_35 for report in book_reports_year),
        'qty_books_other': sum(report.qty_books_other for report in book_reports_year),
        'qty_books_invalid': sum(report.qty_books_invalid for report in book_reports_year),
        'qty_books_out_of_station': sum(report.qty_books_out_of_station for report in book_reports_year),
        'qty_books_neb': sum(report.qty_books_neb for report in book_reports_year),
        'qty_books_prlib': sum(report.qty_books_prlib for report in book_reports_year),
        'qty_books_litres': sum(report.qty_books_litres for report in book_reports_year),
        'qty_books_consultant': sum(report.qty_books_consultant for report in book_reports_year),
        'qty_books_local_library': sum(report.qty_books_local_library for report in book_reports_year),
        'qty_books_part_opl': sum(report.qty_books_part_opl for report in book_reports_year),
        'qty_books_part_enm': sum(report.qty_books_part_enm for report in book_reports_year),
        'qty_books_part_tech': sum(report.qty_books_part_tech for report in book_reports_year),
        'qty_books_part_sh': sum(report.qty_books_part_sh for report in book_reports_year),
        'qty_books_part_si': sum(report.qty_books_part_si for report in book_reports_year),
        'qty_books_part_yl': sum(report.qty_books_part_yl for report in book_reports_year),
        'qty_books_part_hl': sum(report.qty_books_part_hl for report in book_reports_year),
        'qty_books_part_dl': sum(report.qty_books_part_dl for report in book_reports_year),
        'qty_books_part_other': sum(report.qty_books_part_other for report in book_reports_year),
        'qty_books_part_audio': sum(report.qty_books_part_audio for report in book_reports_year),
        'qty_books_part_krai': sum(report.qty_books_part_krai for report in book_reports_year),
        'qty_books_reference_do_14': sum(report.qty_books_reference_do_14 for report in book_reports_year),
        'qty_books_reference_14': sum(report.qty_books_reference_14 for report in book_reports_year),
        'qty_books_reference_18': sum(report.qty_books_reference_18 for report in book_reports_year),
        'qty_books_reference_35': sum(report.qty_books_reference_35 for report in book_reports_year),
        'qty_books_reference_invalid': sum(report.qty_books_reference_invalid for report in book_reports_year),
        'qty_books_reference_online': sum(report.qty_books_reference_online for report in book_reports_year),
    }

    # Записываем данные с начала года в строку 5
    ws['B5'] = (
            year_start_data['qty_books_14'] + year_start_data['qty_books_15_35'] + year_start_data['qty_books_18_35']
            + year_start_data['qty_books_other'] +
            year_start_data['qty_books_neb'] + year_start_data['qty_books_prlib'] +
            year_start_data['qty_books_litres'] + year_start_data['qty_books_consultant'] +
            year_start_data['qty_books_local_library']
    )
    ws['C5'] = year_start_data['qty_books_14']
    ws['D5'] = year_start_data['qty_books_15_35']
    ws['E5'] = year_start_data['qty_books_18_35']
    ws['F5'] = year_start_data['qty_books_other']
    ws['G5'] = year_start_data['qty_books_invalid']
    ws['H5'] = year_start_data['qty_books_out_of_station']
    ws['I5'] = year_start_data['qty_books_neb']
    ws['J5'] = year_start_data['qty_books_prlib']
    ws['K5'] = year_start_data['qty_books_litres']
    ws['L5'] = year_start_data['qty_books_consultant']
    ws['M5'] = year_start_data['qty_books_local_library']
    ws['N5'] = (
            year_start_data['qty_books_part_opl'] + year_start_data['qty_books_part_enm'] +
            year_start_data['qty_books_part_tech'] + year_start_data['qty_books_part_sh'] +
            year_start_data['qty_books_part_si'] + year_start_data['qty_books_part_yl'] +
            year_start_data['qty_books_part_hl'] + year_start_data['qty_books_part_dl'] +
            year_start_data['qty_books_part_other'] + year_start_data['qty_books_part_audio'] +
            year_start_data['qty_books_part_krai']
    )
    ws['O5'] = year_start_data['qty_books_part_opl']
    ws['P5'] = year_start_data['qty_books_part_enm']
    ws['Q5'] = year_start_data['qty_books_part_tech']
    ws['R5'] = year_start_data['qty_books_part_sh']
    ws['S5'] = year_start_data['qty_books_part_si']
    ws['T5'] = year_start_data['qty_books_part_yl']
    ws['U5'] = year_start_data['qty_books_part_hl']
    ws['V5'] = year_start_data['qty_books_part_dl']
    ws['W5'] = year_start_data['qty_books_part_other']
    ws['X5'] = year_start_data['qty_books_part_audio']
    ws['Y5'] = year_start_data['qty_books_part_krai']
    ws['Z5'] = (year_start_data['qty_books_reference_do_14'] + year_start_data['qty_books_reference_14']
                + year_start_data['qty_books_reference_18'] + year_start_data['qty_books_reference_35'])
    ws['AA5'] = year_start_data['qty_books_reference_do_14']
    ws['AB5'] = year_start_data['qty_books_reference_14']
    ws['AC5'] = year_start_data['qty_books_reference_18']
    ws['AD5'] = year_start_data['qty_books_reference_35']
    ws['AE5'] = year_start_data['qty_books_reference_invalid']
    ws['AF5'] = year_start_data['qty_books_reference_online']

    # Заполняем данные за текущий месяц
    current_date = date(year, month, 1)
    month_name_ru = date_format(current_date, "F")
    ws['A1'] = f"Отчет по книговыдаче за {month_name_ru} {year} года"
    ws['V1'] = f"{branch}"

    row_num = 6  # Начинаем с 6 строки (после заголовков)
    for report in book_reports:
        # Вставляем строку перед записью данных (кроме первой строки)
        if row_num > 6:
            ws.insert_rows(row_num)  # Вставляем новую строку
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=col)._style = ws.cell(row=6, column=col)._style  # Копируем стили

        # Записываем данные в текущую строку
        ws[f'A{row_num}'] = report.date.strftime('%d.%m.%Y')
        ws[f'B{row_num}'] = (
                report.qty_books_14 + report.qty_books_15_35 + report.qty_books_18_35 + report.qty_books_other + report.qty_books_neb +
                report.qty_books_prlib + report.qty_books_litres + report.qty_books_consultant + report.qty_books_local_library
        )
        ws[f'C{row_num}'] = report.qty_books_14
        ws[f'D{row_num}'] = report.qty_books_15_35
        ws[f'E{row_num}'] = report.qty_books_18_35
        ws[f'F{row_num}'] = report.qty_books_other
        ws[f'G{row_num}'] = report.qty_books_invalid
        ws[f'H{row_num}'] = report.qty_books_out_of_station
        ws[f'I{row_num}'] = report.qty_books_neb
        ws[f'J{row_num}'] = report.qty_books_prlib
        ws[f'K{row_num}'] = report.qty_books_litres
        ws[f'L{row_num}'] = report.qty_books_consultant
        ws[f'M{row_num}'] = report.qty_books_local_library
        ws[f'N{row_num}'] = (
                report.qty_books_part_opl + report.qty_books_part_enm + report.qty_books_part_tech +
                report.qty_books_part_sh + report.qty_books_part_si + report.qty_books_part_yl +
                report.qty_books_part_hl + report.qty_books_part_dl + report.qty_books_part_other +
                report.qty_books_part_audio + report.qty_books_part_krai
        )
        ws[f'O{row_num}'] = report.qty_books_part_opl
        ws[f'P{row_num}'] = report.qty_books_part_enm
        ws[f'Q{row_num}'] = report.qty_books_part_tech
        ws[f'R{row_num}'] = report.qty_books_part_sh
        ws[f'S{row_num}'] = report.qty_books_part_si
        ws[f'T{row_num}'] = report.qty_books_part_yl
        ws[f'U{row_num}'] = report.qty_books_part_hl
        ws[f'V{row_num}'] = report.qty_books_part_dl
        ws[f'W{row_num}'] = report.qty_books_part_other
        ws[f'X{row_num}'] = report.qty_books_part_audio
        ws[f'Y{row_num}'] = report.qty_books_part_krai
        ws[
            f'Z{row_num}'] = (report.qty_books_reference_do_14 + report.qty_books_reference_14 + report.qty_books_reference_18
                              + report.qty_books_reference_35)
        ws[f'AA{row_num}'] = report.qty_books_reference_do_14
        ws[f'AB{row_num}'] = report.qty_books_reference_14
        ws[f'AC{row_num}'] = report.qty_books_reference_18
        ws[f'AD{row_num}'] = report.qty_books_reference_35
        ws[f'AE{row_num}'] = report.qty_books_reference_invalid
        ws[f'AF{row_num}'] = report.qty_books_reference_online
        ws[f'AG{row_num}'] = report.note
        row_num += 1


def generate_events_report(wb, branch, year, month):
    ws = wb["Массовые мероприятия"]

    # Фильтруем мероприятия по филиалу, году и месяцу
    events = Event.objects.filter(
        library=branch,
        date__year=year,
        date__month=month
    ).order_by('date')

    # Фильтруем мероприятия с начала года до текущего месяца (не включая текущий месяц)
    events_year = Event.objects.filter(
        library=branch,
        date__year=year,
        date__month__lt=month
    ).order_by('date')

    # Агрегируем данные с начала года
    year_start_data = {
        'quantity': sum(event.quantity for event in events_year),
        'participants': sum(event.age_14 + event.age_18 + event.age_35 + event.age_other for event in events_year),
        'age_14': sum(event.age_14 for event in events_year),
        'age_18': sum(event.age_18 for event in events_year),
        'age_35': sum(event.age_35 for event in events_year),
        'age_other': sum(event.age_other for event in events_year),
        'invalids': sum(event.invalids for event in events_year),
        'pensioners': sum(event.pensioners for event in events_year),
        'out_of_station': sum(event.out_of_station for event in events_year),
    }

    # Записываем данные с начала года в строку 6
    ws['D5'] = year_start_data['quantity']
    ws['E5'] = year_start_data['participants']
    ws['F5'] = year_start_data['age_14']
    ws['G5'] = year_start_data['age_18']
    ws['H5'] = year_start_data['age_35']
    ws['I5'] = year_start_data['age_other']
    ws['J5'] = year_start_data['invalids']
    ws['K5'] = year_start_data['pensioners']
    ws['L5'] = year_start_data['out_of_station']

    current_date = date(year, month, 1)
    month_name_ru = date_format(current_date, "F")
    # Заполняем данные за текущий месяц
    ws['A1'] = f"Отчет по мероприятиям за {month_name_ru} {year} года"
    ws['G1'] = f"{branch}"

    row_num = 6  # Начинаем с 7 строки (после заголовков)
    for event in events:
        # Вставляем строку перед записью данных (кроме первой строки)
        if row_num > 6:
            ws.insert_rows(row_num)  # Вставляем новую строку
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=col)._style = ws.cell(row=6, column=col)._style  # Копируем стили

        # Записываем данные в текущую строку
        ws[f'A{row_num}'] = event.date.strftime('%d.%m.%Y')
        ws[f'B{row_num}'] = event.name
        ws[f'C{row_num}'] = event.direction
        ws[f'D{row_num}'] = event.quantity
        ws[f'E{row_num}'] = event.age_14 + event.age_18 + event.age_35 + event.age_other
        ws[f'F{row_num}'] = event.age_14
        ws[f'G{row_num}'] = event.age_18
        ws[f'H{row_num}'] = event.age_35
        ws[f'I{row_num}'] = event.age_other
        ws[f'J{row_num}'] = event.invalids
        ws[f'K{row_num}'] = event.pensioners
        ws[f'L{row_num}'] = event.out_of_station
        ws[f'M{row_num}'] = event.as_part
        ws[f'N{row_num}'] = 'X' if event.paid else ''
        ws[f'O{row_num}'] = event.note
        row_num += 1


def generate_quarter_excel(user, year, quarter):
    try:
        employee = Employee.objects.get(user=user)
        branch = employee.branch
    except Employee.DoesNotExist:
        return None

    if not branch:
        return None

    # Устанавливаем русскую локаль только для этого потока
    old_locale = locale.getlocale()
    try:
        locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
    except locale.Error:
        # Если не удалось установить локаль, используем fallback
        pass

    # Загрузка шаблона
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(base_dir, 'reports/excell/report_quarter_template.xlsx')

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"File not found: {template_path}")

    wb = load_workbook(filename=template_path)
    ws = wb.active

    # Динамическое заполнение заголовков
    ws['E3'] = f"Значение показателя в {quarter} квартале {year} года"
    ws['C3'] = f"Данные статистики по посещаемости по итогам {year - 1} года"
    ws['D3'] = f"Значение целевого показателя за {year} год план"
    ws['W3'] = f"итого за {quarter}-й квартал {year}"

    # Полное наименование филиала
    ws['B7'] = branch.full_name

    # Данные за предыдущий год
    previous_year = year - 1
    total_prev_year = get_total_previous_year(branch, previous_year)
    ws['C7'] = total_prev_year

    # Плановые показатели за текущий год
    total_plan = get_total_plan(branch, year)
    ws['D7'] = total_plan

    # Данные за выбранный квартал
    months = {
        1: (1, 2, 3),
        2: (4, 5, 6),
        3: (7, 8, 9),
        4: (10, 11, 12)
    }[quarter]

    # Заполнение данных по месяцам
    for i, month in enumerate(months):
        col_offset = i * 6
        fill_month_data(ws, branch, year, month, col_offset)

    # Итог за квартал
    total_quarter = 0
    for i, month in enumerate(months):
        col_offset = i * 6
        total_quarter += (ws[f'{chr(74 + col_offset)}7'].value or 0)

    ws['W7'] = total_quarter

    # Итог за год
    total_year = get_total_year(branch, year)
    ws['X7'] = total_year

    # Процент выполнения плана
    ws['Y7'] = (ws['X7'].value / (ws['D7'].value or 1)) * 100 if ws['X7'].value is not None else 0

    # Динамическое заполнение названий месяцев
    month_names = []
    for month in months:
        current_date = datetime(year, month, 1)
        month_name_ru = date_format(current_date, "F")
        month_names.append(month_name_ru)

    ws['E4'] = f"{month_names[0]} посещаемость (всего)"
    if len(months) > 1:
        ws['K4'] = f"{month_names[1]} посещаемость (всего)"
    if len(months) > 2:
        ws['Q4'] = f"{month_names[2]} посещаемость (всего)"

    # Возвращаем язык по умолчанию
    translation.deactivate()

    # Сохранение файла
    filename = f"квартальный_нац_цели_{branch.short_name}_{year}_Q{quarter}.xlsx"
    safe_filename = quote(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
    wb.save(response)
    return response


def get_total_previous_year(branch, previous_year):
    visit_reports_prev_year = VisitReport.objects.filter(library=branch, date__year=previous_year)
    events_prev_year = Event.objects.filter(library=branch, date__year=previous_year)
    book_reports_prev_year = BookReport.objects.filter(library=branch, date__year=previous_year)

    total_visits_prev_year = sum(
        (report.qty_visited_14 or 0) + (report.qty_visited_15_35 or 0) + (report.qty_visited_18_35 or 0)
        + (report.qty_visited_other or 0)
        for report in visit_reports_prev_year
    )

    total_events_prev_year = sum(
        (event.age_14 or 0) + (event.age_18 or 0) + (event.age_35 or 0) + (event.age_other or 0)
        for event in events_prev_year
    )

    total_online_prev_year = sum(
        (event.online or 0) for event in events_prev_year
    ) + sum(
        (report.qty_visited_online or 0) + (report.qty_visited_prlib or 0) + (report.qty_visited_litres or 0)
        for report in visit_reports_prev_year
    ) + sum(
        (report.qty_books_reference_online or 0)
        for report in book_reports_prev_year
    )

    total_out_station_prev_year = sum(
        (event.out_of_station or 0) for event in events_prev_year
    ) + sum(
        (report.qty_visited_out_station or 0)
        for report in visit_reports_prev_year
    )

    total_prev_year = (
            total_visits_prev_year +
            total_events_prev_year +
            total_online_prev_year +
            total_out_station_prev_year
    )

    if not visit_reports_prev_year.exists() and not events_prev_year.exists() and not book_reports_prev_year.exists():
        visits_first_data = VisitsFirstData.objects.filter(library=branch).first()
        events_first_data = EventsFirstData.objects.filter(library=branch).first()
        books_first_data = BooksFirstData.objects.filter(
            library=branch).first()  # Если есть аналогичная модель для книг
        total_prev_year = ((visits_first_data.data or 0) if visits_first_data else 0) + (
            (events_first_data.data or 0) if events_first_data else 0) + (
                              (books_first_data.data or 0) if books_first_data else 0)

    return total_prev_year


def get_total_plan(branch, year):
    visit_plan = VisitPlan.objects.filter(library=branch, year=year).first()
    events_plan = EventsPlan.objects.filter(library=branch, year=year).first()
    total_plan = ((visit_plan.total_visits or 0) if visit_plan else 0) + (
        (events_plan.total_events or 0) if events_plan else 0)
    return total_plan


def fill_month_data(ws, branch, year, month, col_offset):
    visit_reports = VisitReport.objects.filter(library=branch, date__year=year, date__month=month)
    events = Event.objects.filter(library=branch, date__year=year, date__month=month)
    book_reports = BookReport.objects.filter(library=branch, date__year=year, date__month=month)

    total_visits = sum(
        (report.qty_visited_14 or 0) + (report.qty_visited_15_35 or 0) + (report.qty_visited_18_35 or 0)
        + (report.qty_visited_other or 0)
        for report in visit_reports
    )
    total_events = sum(
        ((event.age_14 or 0) + (event.age_18 or 0) + (event.age_35 or 0) + (event.age_other or 0)) - (event.out_of_station or 0)
        for event in events
    )
    total_online = sum(
        (event.online or 0) for event in events
    ) + sum(
        (report.qty_visited_online or 0) + (report.qty_visited_prlib or 0) + (report.qty_visited_litres or 0)
        for report in visit_reports
    ) + sum(
        (report.qty_books_reference_online or 0)
        for report in book_reports
    )
    total_out_station = sum(
        (event.out_of_station or 0) for event in events
    ) + sum(
        (report.qty_visited_out_station or 0)
        for report in visit_reports
    )
    total_events_out_station = sum(
        (event.out_of_station or 0) for event in events
    )

    # Записываем данные в ячейки
    ws[f'{chr(69 + col_offset)}7'] = total_visits + total_events
    ws[f'{chr(70 + col_offset)}7'] = total_events
    ws[f'{chr(71 + col_offset)}7'] = total_online
    ws[f'{chr(72 + col_offset)}7'] = total_out_station
    ws[f'{chr(73 + col_offset)}7'] = total_events_out_station
    ws[f'{chr(74 + col_offset)}7'] = (ws[f'{chr(69 + col_offset)}7'].value or 0) + (
            ws[f'{chr(71 + col_offset)}7'].value or 0) + (ws[f'{chr(72 + col_offset)}7'].value or 0)


def get_total_year(branch, year):
    visit_reports_year = VisitReport.objects.filter(library=branch, date__year=year)
    events_year = Event.objects.filter(library=branch, date__year=year)
    book_reports_year = BookReport.objects.filter(library=branch, date__year=year)

    total_year = (
            sum((report.qty_visited_14 or 0) + (report.qty_visited_15_35 or 0) + (report.qty_visited_18_35 or 0)
                + (report.qty_visited_other or 0) + ( report.qty_visited_out_station or 0) for report in visit_reports_year) +
            sum((event.age_14 or 0) + (event.age_18 or 0) + (event.age_35 or 0) + (event.age_other or 0) for event in events_year) +
            sum((event.online or 0) for event in events_year) +
            sum((report.qty_visited_online or 0) + (report.qty_visited_prlib or 0) + (report.qty_visited_litres or 0)
                for report in visit_reports_year) +
            sum((report.qty_books_reference_online or 0) for report in book_reports_year)
    )
    return total_year

    # Процент выполнения плана
    ws['Y7'] = ws['X7'].value / (ws['D7'].value / 100) if ws['D7'].value else 0

    # Получаем названия месяцев
    month_names = [format_datetime(datetime(year, month, 1), 'MMMM', locale='ru') for month in months]

    # Записываем названия месяцев в ячейки Excel
    ws['E4'] = f"{month_names[1]} посещаемость (всего)" if len(month_names) >= 1 else ""
    ws['K4'] = f"{month_names[2]} посещаемость (всего)" if len(month_names) >= 2 else ""
    ws['Q4'] = f"{month_names[3]} посещаемость (всего)" if len(month_names) >= 3 else ""

    # Сохранение файла
    filename = f"квартальный_нац_цели_{branch.short_name}_{year}_Q{quarter}.xlsx"
    safe_filename = quote(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
    wb.save(response)
    return response


def generate_digital_month_report(user, month):
    try:
        employee = Employee.objects.get(user=user)
        branch = employee.branch
    except Employee.DoesNotExist:
        return None

    if not branch:
        return None

    current_year = datetime.now().year

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(base_dir, 'reports/excell/digital_month_template.xlsx')

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"File not found: {template_path}")

    wb = load_workbook(filename=template_path)
    ws = wb.active

    current_date = date(current_year, month, 1)
    month_name = date_format(current_date, "F")
    ws['A1'] = f"Цифровой отчёт о работе {branch.full_name} (МБУК «ЦБС им. А. Белого» Балашиха) за {month_name} {current_year} года (с нарастающим итогом)"

    visit_reports = VisitReport.objects.filter(library=branch, date__year=current_year, date__month__lte=month)
    book_reports = BookReport.objects.filter(library=branch, date__year=current_year, date__month__lte=month)
    events = Event.objects.filter(library=branch, date__year=current_year, date__month__lte=month)

    regs_plan = RegsPlan.objects.filter(library=branch, year=current_year).first()
    book_plan = BookPlan.objects.filter(library=branch, year=current_year).first()
    visit_plan = VisitPlan.objects.filter(library=branch, year=current_year).first()

    # === Пользователи (B4–C10) ===
    ws['B4'] = regs_plan.total_regs if regs_plan else 0
    total_regs = sum(
        (r.qty_reg_14 or 0) +
        (r.qty_reg_15_35 or 0) +      # 15–17
        (r.qty_reg_18_35 or 0) +      # 18–35
        (r.qty_reg_other or 0) +
        (r.qty_reg_prlib or 0) +
        (r.qty_reg_litres or 0) +
        (r.qty_reg_out_of_station or 0)
        for r in visit_reports
    )
    ws['C4'] = total_regs
    ws['D4'] = (ws['C4'].value / ws['B4'].value) * 100 if ws['B4'].value else 0

    ws['C5'] = sum(r.qty_reg_14 or 0 for r in visit_reports)
    ws['C6'] = sum(r.qty_reg_15_35 or 0 for r in visit_reports)   # 15–17
    ws['C7'] = sum(r.qty_reg_18_35 or 0 for r in visit_reports)   # 18–35
    ws['C8'] = sum(
        (r.qty_reg_14 or 0) + (r.qty_reg_15_35 or 0) + (r.qty_reg_18_35 or 0) +
        (r.qty_reg_other or 0) + (r.qty_reg_invalid or 0) + (r.qty_reg_pensioners or 0)
        for r in visit_reports
    )  # стационар
    ws['C9'] = sum(r.qty_reg_out_of_station or 0 for r in visit_reports)  # вне стационара
    ws['C10'] = sum((r.qty_reg_prlib or 0) + (r.qty_reg_litres or 0) for r in visit_reports)  # удалённые

    # === Книговыдача (B11–C17) ===
    ws['B11'] = book_plan.total_books if book_plan else 0
    total_books = sum(
        (r.qty_books_14 or 0) +
        (r.qty_books_15_35 or 0) +
        (r.qty_books_18_35 or 0) +
        (r.qty_books_other or 0) +
        (r.qty_books_out_of_station or 0) +
        (r.qty_books_neb or 0) +
        (r.qty_books_prlib or 0) +
        (r.qty_books_litres or 0) +
        (r.qty_books_consultant or 0) +
        (r.qty_books_local_library or 0)
        for r in book_reports
    )
    ws['C11'] = total_books
    ws['D11'] = (ws['C11'].value / ws['B11'].value) * 100 if ws['B11'].value else 0

    ws['C12'] = sum(r.qty_books_14 or 0 for r in book_reports)
    ws['C13'] = sum(r.qty_books_15_35 or 0 for r in book_reports)   # 15–17
    ws['C14'] = sum(r.qty_books_18_35 or 0 for r in book_reports)   # 18–35
    ws['C15'] = sum(
        (r.qty_books_14 or 0) + (r.qty_books_15_35 or 0) + (r.qty_books_18_35 or 0) +
        (r.qty_books_other or 0) + (r.qty_books_invalid or 0) +
        (r.qty_books_neb or 0) + (r.qty_books_prlib or 0) + (r.qty_books_litres or 0) +
        (r.qty_books_consultant or 0) + (r.qty_books_local_library or 0) +
        (r.qty_books_part_opl or 0) + (r.qty_books_part_enm or 0) + (r.qty_books_part_tech or 0) +
        (r.qty_books_part_sh or 0) + (r.qty_books_part_si or 0) + (r.qty_books_part_yl or 0) +
        (r.qty_books_part_hl or 0) + (r.qty_books_part_dl or 0) + (r.qty_books_part_other or 0) +
        (r.qty_books_part_audio or 0) + (r.qty_books_part_krai or 0)
        for r in book_reports
    ) - sum(r.qty_books_out_of_station or 0 for r in book_reports)  # приблизительно — стационар
    ws['C16'] = sum(r.qty_books_out_of_station or 0 for r in book_reports)  # вне стационара
    ws['C17'] = sum(r.qty_books_prlib or 0 + r.qty_books_litres or 0 for r in book_reports)  # удалённые

    # === Общая посещаемость (B18–C24) ===
    visits_from_reports = sum(
        (r.qty_visited_14 or 0) +
        (r.qty_visited_15_35 or 0) +
        (r.qty_visited_18_35 or 0) +
        (r.qty_visited_other or 0) +
        (r.qty_visited_invalids or 0) +
        (r.qty_visited_pensioners or 0) +
        (r.qty_visited_out_station or 0) +
        (r.qty_visited_online or 0) +
        (r.qty_visited_prlib or 0) +
        (r.qty_visited_litres or 0)
        for r in visit_reports
    )

    event_attendees = sum(
        (e.age_14 or 0) +
        (e.age_18 or 0) +    # 15–17
        (e.age_35 or 0) +    # 18–35
        (e.age_other or 0) +
        (e.invalids or 0) +
        (e.pensioners or 0)
        for e in events
    )

    references_total = sum(
        (r.qty_books_reference_do_14 or 0) +
        (r.qty_books_reference_14 or 0) +
        (r.qty_books_reference_18 or 0) +  # 15–17
        (r.qty_books_reference_35 or 0) +  # 18–35
        (r.qty_books_reference_other or 0) +
        (r.qty_books_reference_invalid or 0) +
        (r.qty_books_reference_online or 0)
        for r in book_reports
    )

    ws['B18'] = visit_plan.total_visits if visit_plan else 0
    ws['C18'] = visits_from_reports + event_attendees + references_total
    ws['D18'] = (ws['C18'].value / ws['B18'].value) * 100 if ws['B18'].value else 0

    ws['C19'] = sum(r.qty_visited_14 or 0 for r in visit_reports) + sum(e.age_14 or 0 for e in events)
    ws['C20'] = sum(r.qty_visited_15_35 or 0 for r in visit_reports) + sum(e.age_18 or 0 for e in events)  # 15–17
    ws['C21'] = sum(r.qty_visited_18_35 or 0 for r in visit_reports) + sum(e.age_35 or 0 for e in events)  # 18–35
    ws['C22'] = sum(
        (r.qty_visited_14 or 0) + (r.qty_visited_15_35 or 0) + (r.qty_visited_18_35 or 0) +
        (r.qty_visited_other or 0)
        for r in visit_reports
    ) + sum(
        (e.age_14 or 0) + (e.age_18 or 0) + (e.age_35 or 0) + (e.age_other or 0) +
        (e.invalids or 0) + (e.pensioners or 0)
        for e in events if e.out_of_station == 0
    )  # стационар
    ws['C23'] = sum(r.qty_visited_out_station or 0 for r in visit_reports) + sum(
        (e.age_14 or 0) + (e.age_18 or 0) + (e.age_35 or 0) + (e.age_other or 0)
        for e in events if e.out_of_station > 0
    )  # вне стационара
    ws['C24'] = sum(r.qty_visited_online or 0 + r.qty_visited_prlib or 0 + r.qty_visited_litres or 0 for r in visit_reports) + sum(r.qty_books_reference_online or 0 for r in book_reports)  # удалённые

    # === Инвалиды, платные, льготы (C25–C27) ===
    ws['C25'] = sum(r.qty_visited_invalids or 0 for r in visit_reports) + sum(e.invalids or 0 for e in events)
    ws['C26'] = sum((e.age_14 or 0) + (e.age_18 or 0) + (e.age_35 or 0) + (e.age_other or 0) for e in events if e.paid)
    ws['C27'] = sum(e.invalids or 0 for e in events if e.paid)

    # === Справки (C28–C32) ===
    ws['C28'] = sum(
        (r.qty_books_reference_do_14 or 0) +
        (r.qty_books_reference_14 or 0) +
        (r.qty_books_reference_18 or 0) +
        (r.qty_books_reference_35 or 0) +
        (r.qty_books_reference_other or 0) +
        (r.qty_books_reference_invalid or 0) +
        (r.qty_books_reference_online or 0)
        for r in book_reports
    )
    ws['C29'] = sum(r.qty_books_reference_do_14 or 0 for r in book_reports)
    ws['C30'] = sum(r.qty_books_reference_14 or 0 for r in book_reports)      # 15–17
    ws['C31'] = sum(r.qty_books_reference_18 or 0 for r in book_reports)      # 18–35
    ws['C32'] = sum(r.qty_books_reference_online or 0 for r in book_reports)

    # === Мероприятия (всего) (C33–C40) ===
    ws['C33'] = sum(e.quantity for e in events)
    ws['C34'] = sum((e.age_14 or 0) + (e.age_18 or 0) + (e.age_35 or 0) + (e.age_other or 0) for e in events)
    # До 14
    ev_u14 = [e for e in events if e.age_14 > 0]
    ws['C35'] = sum(e.quantity for e in ev_u14)
    ws['C36'] = sum(e.age_14 for e in ev_u14)
    # 15–17
    ev_15_17 = [e for e in events if e.age_18 > 0]
    ws['C37'] = sum(e.quantity for e in ev_15_17)
    ws['C38'] = sum(e.age_18 for e in ev_15_17)
    # 18–35
    ev_18_35 = [e for e in events if e.age_35 > 0]
    ws['C39'] = sum(e.quantity for e in ev_18_35)
    ws['C40'] = sum(e.age_35 for e in ev_18_35)

    # === Мероприятия в стационаре (C41–C48) ===
    stat_events = [e for e in events if e.out_of_station == 0]
    ws['C41'] = sum(e.quantity for e in stat_events)
    ws['C42'] = sum((e.age_14 or 0) + (e.age_18 or 0) + (e.age_35 or 0) + (e.age_other or 0) for e in stat_events)
    # До 14
    s_u14 = [e for e in stat_events if e.age_14 > 0]
    ws['C43'] = sum(e.quantity for e in s_u14)
    ws['C44'] = sum(e.age_14 for e in s_u14)
    # 15–17
    s_15_17 = [e for e in stat_events if e.age_18 > 0]
    ws['C45'] = sum(e.quantity for e in s_15_17)
    ws['C46'] = sum(e.age_18 for e in s_15_17)
    # 18–35
    s_18_35 = [e for e in stat_events if e.age_35 > 0]
    ws['C47'] = sum(e.quantity for e in s_18_35)
    ws['C48'] = sum(e.age_35 for e in s_18_35)

    # === Мероприятия вне стационара (C49–C50) ===
    out_stat_events = [e for e in events if e.out_of_station > 0]
    ws['C49'] = sum(e.quantity for e in out_stat_events)
    ws['C50'] = sum((e.age_14 or 0) + (e.age_18 or 0) + (e.age_35 or 0) + (e.age_other or 0) for e in out_stat_events)

    # Сохранение
    filename = f"цифровой_ежемесячный_{branch.short_name}_{current_year}_{month}.xlsx"
    safe_filename = quote(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
    wb.save(response)
    return response


def generate_nats_project_report(user, year, month):
    try:
        employee = Employee.objects.get(user=user)
        branch = employee.branch
    except Employee.DoesNotExist:
        return None

    if not branch:
        return None

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(base_dir, 'reports/excell/nats_project_template.xlsx')

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"File not found: {template_path}")

    wb = load_workbook(filename=template_path)
    ws = wb.active

    # Заголовки
    ws['I3'] = f"{year} год"
    regs_plan = RegsPlan.objects.filter(library=branch, year=year).first()
    visit_plan = VisitPlan.objects.filter(library=branch, year=year).first()

    ws['B5'] = f"Кол-во пользователей библиотек. Регистрация читателей План на {year} год {regs_plan.total_regs if regs_plan else 0} чел."
    ws['G5'] = f"Общая посещаемость (план на {year} г. {visit_plan.total_visits if visit_plan else 0}) Факт Число посещений библиотек ВСЕГО (нарастающий итог с начала года)"
    ws['I6'] = f"Всего в {year} году"
    ws['J6'] = f"В том числе в рамках Нацпроекта 'Творческие люди в {year} году'"

    # Данные
    visit_reports = VisitReport.objects.filter(library=branch, date__year=year, date__month__lte=month)
    book_reports = BookReport.objects.filter(library=branch, date__year=year, date__month__lte=month)
    events = Event.objects.filter(library=branch, date__year=year, date__month__lte=month)

    # A7 — название
    ws['A7'] = branch.full_name

    # B7 — дети до 14 лет (регистрации)
    reg_u14 = sum(r.qty_reg_14 or 0 for r in visit_reports)
    ws['B7'] = reg_u14

    # C7 — остальное население (15+): все регистрации, кроме до 14
    reg_15_plus = (
        sum(r.qty_reg_15_35 or 0 for r in visit_reports) +      # 15–17
        sum(r.qty_reg_18_35 or 0 for r in visit_reports) +      # 18–35
        sum(r.qty_reg_other or 0 for r in visit_reports)    # 35+
    )
    ws['C7'] = reg_15_plus

    # D7 — ВСЕГО = B7 + C7
    ws['D7'] = reg_u14 + reg_15_plus

    # E7 — количество мероприятий (штук)
    ws['E7'] = sum(e.quantity or 0 for e in events)

    # F7 — участники мероприятий (люди)
    event_attendees = (
        sum(e.age_14 or 0 for e in events) +    # до 14
        sum(e.age_18 or 0 for e in events) +    # 15–17 ← ИСПРАВЛЕНО!
        sum(e.age_35 or 0 for e in events) +    # 18–35
        sum(e.age_other or 0 for e in events)   # 35+
    )
    ws['F7'] = event_attendees

    # G7 — общая посещаемость = посещения + мероприятия + справки
    visits = (
        sum(r.qty_visited_14 or 0 for r in visit_reports) +
        sum(r.qty_visited_15_35 or 0 for r in visit_reports) +
        sum(r.qty_visited_18_35 or 0 for r in visit_reports) +
        sum(r.qty_visited_other or 0 for r in visit_reports) +
        sum(r.qty_visited_out_station or 0 for r in visit_reports) +
        sum(r.qty_visited_online or 0 for r in visit_reports) +
        sum(r.qty_visited_prlib or 0 for r in visit_reports) +
        sum(r.qty_visited_litres or 0 for r in visit_reports)
    )
    references = sum(r.qty_books_reference_online or 0 for r in book_reports)
    ws['G7'] = visits + event_attendees + references

    # H7 — выданные книги
    books = (
        sum(r.qty_books_14 or 0 for r in book_reports) +
        sum(r.qty_books_15_35 or 0 for r in book_reports) +
        sum(r.qty_books_18_35 or 0 for r in book_reports) +
        sum(r.qty_books_other or 0 for r in book_reports) +
        sum(r.qty_books_out_of_station or 0 for r in book_reports) +
        sum(r.qty_books_neb or 0 for r in book_reports) +
        sum(r.qty_books_prlib or 0 for r in book_reports) +
        sum(r.qty_books_litres or 0 for r in book_reports) +
        sum(r.qty_books_consultant or 0 for r in book_reports) +
        sum(r.qty_books_local_library or 0 for r in book_reports)
        # остальные поля (разделы литературы) — можно добавить при необходимости
    )
    ws['H7'] = books

    # K7 — посетители сайтов (удалённые обращения)
    # Согласно шаблону — это онлайн-обращения, а не онлайн-мероприятия
    ws['K7'] = sum(r.qty_visited_online or 0 for r in visit_reports)

    # Сохранение
    filename = f"нац_проект_{branch.short_name}_{year}_{month}.xlsx"
    safe_filename = quote(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
    wb.save(response)
    return response


def generate_quarter_excel_all_branches(user, year, quarter):
    try:
        employee = Employee.objects.get(user=user)
    except Employee.DoesNotExist:
        return None

    # Устанавливаем русскую локаль только для этого потока
    old_locale = locale.getlocale()
    try:
        locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
    except locale.Error:
        pass

    # Загрузка нового шаблона
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(base_dir, 'reports/excell/report_quarter_template_all_branch.xlsx')

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"File not found: {template_path}")

    wb = load_workbook(filename=template_path)
    ws = wb.active

    # Динамическое заполнение заголовков
    ws['E3'] = f"Значение показателя в {quarter} квартале {year} года"
    ws['C3'] = f"Данные статистики по посещаемости по итогам {year - 1} года"
    ws['D3'] = f"Значение целевого показателя за {year} год план"
    ws['W3'] = f"итого за {quarter}-й квартал {year}"

    # Получаем все филиалы, кроме "Администрация", в том же порядке, что и в шаблоне
    branches = Branch.objects.exclude(department=True).order_by('id')

    # Проверяем, что филиалов ровно 17, как в шаблоне
    if branches.count() != 17:
        raise ValueError("Количество филиалов не соответствует шаблону (ожидается 17)")

    # Определяем месяцы квартала
    months = {
        1: (1, 2, 3),
        2: (4, 5, 6),
        3: (7, 8, 9),
        4: (10, 11, 12)
    }[quarter]

    # Начальная и конечная строки для данных
    start_row = 7
    end_row = 23
    current_row = start_row

    # Для хранения итоговых значений
    totals = {
        'prev_year': 0,
        'plan': 0,
        'quarter': 0,
        'year': 0,
        'months': {month: {
            'visits_events': 0,
            'events': 0,
            'online': 0,
            'out_station': 0,
            'events_out_station': 0,
            'total': 0
        } for month in months}
    }

    for branch in branches:
        # Заполняем название филиала (колонка B)
        ws[f'B{current_row}'] = branch.full_name

        # Данные за предыдущий год
        total_prev_year = get_total_previous_year(branch, year - 1)
        ws[f'C{current_row}'] = total_prev_year
        totals['prev_year'] += total_prev_year

        # Плановые показатели за текущий год
        total_plan = get_total_plan(branch, year)
        ws[f'D{current_row}'] = total_plan
        totals['plan'] += total_plan

        # Заполнение данных по месяцам
        for i, month in enumerate(months):
            col_offset = i * 6
            month_data = fill_month_data_all_branches(ws, branch, year, month, col_offset, current_row)

            # Суммируем данные для итогов
            for key in month_data:
                totals['months'][month][key] += month_data[key]

        # Итог за квартал
        total_quarter = 0
        for i, month in enumerate(months):
            col_offset = i * 6
            total_quarter += (ws[f'{chr(74 + col_offset)}{current_row}'].value or 0)

        ws[f'W{current_row}'] = total_quarter
        totals['quarter'] += total_quarter

        # Итог за год
        total_year = get_total_year(branch, year)
        ws[f'X{current_row}'] = total_year
        totals['year'] += total_year

        # Процент выполнения плана
        percent = (total_year / (total_plan or 1)) * 100 if total_year is not None else 0
        ws[f'Y{current_row}'] = round(percent, 2)  # Округляем до 2 знаков после запятой

        current_row += 1
        if current_row > end_row:
            break  # На случай, если филиалов больше 17

    # Заполняем строку с итогами (строка 24 в новом шаблоне)
    current_row = 24
    ws[f'A{current_row}'] = "ИТОГО:"
    ws[f'C{current_row}'] = totals['prev_year']
    ws[f'D{current_row}'] = totals['plan']

    # Заполняем итоги по месяцам
    for i, month in enumerate(months):
        col_offset = i * 6
        month_data = totals['months'][month]

        ws[f'{chr(69 + col_offset)}{current_row}'] = month_data['visits_events']
        ws[f'{chr(70 + col_offset)}{current_row}'] = month_data['events']
        ws[f'{chr(71 + col_offset)}{current_row}'] = month_data['online']
        ws[f'{chr(72 + col_offset)}{current_row}'] = month_data['out_station']
        ws[f'{chr(73 + col_offset)}{current_row}'] = month_data['events_out_station']
        ws[f'{chr(74 + col_offset)}{current_row}'] = month_data['total']

    ws[f'W{current_row}'] = totals['quarter']
    ws[f'X{current_row}'] = totals['year']

    total_percent = (totals['year'] / (totals['plan'] or 1)) * 100 if totals['year'] is not None else 0
    ws[f'Y{current_row}'] = round(total_percent, 2)  # Округляем до 2 знаков после запятой

    # Динамическое заполнение названий месяцев
    month_names = []
    for month in months:
        current_date = datetime(year, month, 1)
        month_name_ru = date_format(current_date, "F")
        month_names.append(month_name_ru)

    ws['E4'] = f"{month_names[0]} посещаемость (всего)"
    if len(months) > 1:
        ws['K4'] = f"{month_names[1]} посещаемость (всего)"
    if len(months) > 2:
        ws['Q4'] = f"{month_names[2]} посещаемость (всего)"

    # Сохранение файла
    filename = f"квартальный_нац_цели_все_филиалы_{year}_Q{quarter}.xlsx"
    safe_filename = quote(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
    wb.save(response)

    # Восстанавливаем локаль
    locale.setlocale(locale.LC_ALL, old_locale)

    return response


def fill_month_data_all_branches(ws, branch, year, month, col_offset, row):
    visit_reports = VisitReport.objects.filter(library=branch, date__year=year, date__month=month)
    events = Event.objects.filter(library=branch, date__year=year, date__month=month)
    book_reports = BookReport.objects.filter(library=branch, date__year=year, date__month=month)

    total_visits = sum(
        (report.qty_visited_14 or 0) + (report.qty_visited_15_35 or 0) + (report.qty_visited_other or 0)
        for report in visit_reports
    )
    total_events = sum(
        ((event.age_14 or 0) + (event.age_35 or 0) + (event.age_other or 0)) - (event.out_of_station or 0)
        for event in events
    )
    total_online = sum(
        (event.online or 0) for event in events
    ) + sum(
        (report.qty_visited_online or 0) + (report.qty_visited_prlib or 0) + (report.qty_visited_litres or 0)
        for report in visit_reports
    ) + sum(
        (report.qty_books_reference_online or 0)
        for report in book_reports
    )
    total_out_station = sum(
        (event.out_of_station or 0) for event in events
    ) + sum(
        (report.qty_visited_out_station or 0)
        for report in visit_reports
    )
    total_events_out_station = sum(
        (event.out_of_station or 0) for event in events
    )
    total_month = total_visits + total_events + total_online + total_out_station

    # Записываем данные в ячейки
    ws[f'{chr(69 + col_offset)}{row}'] = total_visits + total_events
    ws[f'{chr(70 + col_offset)}{row}'] = total_events
    ws[f'{chr(71 + col_offset)}{row}'] = total_online
    ws[f'{chr(72 + col_offset)}{row}'] = total_out_station
    ws[f'{chr(73 + col_offset)}{row}'] = total_events_out_station
    ws[f'{chr(74 + col_offset)}{row}'] = total_month

    return {
        'visits_events': total_visits + total_events,
        'events': total_events,
        'online': total_online,
        'out_station': total_out_station,
        'events_out_station': total_events_out_station,
        'total': total_month
    }


def generate_nats_project_report_all_branches(user, year, month):
    try:
        employee = Employee.objects.get(user=user)
    except Employee.DoesNotExist:
        return None

    # Загрузка шаблона для всех филиалов
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(base_dir, 'reports/excell/nats_project_template_all_branches.xlsx')

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"File not found: {template_path}")

    wb = load_workbook(filename=template_path)
    ws = wb['2025_год']  # Работаем с нужным листом

    # Заполнение заголовков
    ws['I3'] = f"{year} год"

    # Получаем все филиалы, исключая административные (department=True)
    branches = Branch.objects.exclude(department=True).order_by('id')

    # Начальная строка для данных (в шаблоне данные начинаются с 8 строки)
    start_row = 7
    current_row = start_row

    # Для хранения итоговых значений
    totals = {
        'children': 0,
        'adults': 0,
        'total_regs': 0,
        'events_count': 0,
        'events_participants': 0,
        'total_visits': 0,
        'books_issued': 0,
        'website_visits': 0
    }

    for branch in branches:
        # Получаем данные с начала года до конца выбранного месяца
        visit_reports = VisitReport.objects.filter(library=branch, date__year=year, date__month__lte=month)
        book_reports = BookReport.objects.filter(library=branch, date__year=year, date__month__lte=month)
        events = Event.objects.filter(library=branch, date__year=year, date__month__lte=month)

        # Рассчитываем показатели
        children = sum(report.qty_reg_14 or 0 for report in visit_reports)
        adults = sum(report.qty_reg_15_35 or 0 for report in visit_reports) + \
                 sum(report.qty_reg_other or 0 for report in visit_reports) + \
                 sum(report.qty_reg_out_of_station or 0 for report in visit_reports) + \
                 sum(report.qty_reg_prlib or 0 for report in visit_reports) + \
                 sum(report.qty_reg_litres or 0 for report in visit_reports)
        total_regs = children + adults
        events_count = sum(event.quantity or 0 for event in events)
        events_participants = sum(event.age_14 or 0 for event in events) + \
                              sum(event.age_35 or 0 for event in events) + \
                              sum(event.age_other or 0 for event in events)
        total_visits = events_participants + \
                       sum(report.qty_visited_14 or 0 for report in visit_reports) + \
                       sum(report.qty_visited_15_35 or 0 for report in visit_reports) + \
                       sum(report.qty_visited_other or 0 for report in visit_reports) + \
                       sum(report.qty_visited_out_station or 0 for report in visit_reports) + \
                       sum(report.qty_visited_online or 0 for report in visit_reports) + \
                       sum(report.qty_visited_prlib or 0 for report in visit_reports) + \
                       sum(report.qty_visited_litres or 0 for report in visit_reports) + \
                       sum(report.qty_books_reference_online or 0 for report in book_reports)
        books_issued = sum(report.qty_books_14 or 0 for report in book_reports) + \
                       sum(report.qty_books_15_35 or 0 for report in book_reports) + \
                       sum(report.qty_books_other or 0 for report in book_reports) + \
                       sum(report.qty_books_out_of_station or 0 for report in book_reports) + \
                       sum(report.qty_books_neb or 0 for report in book_reports) + \
                       sum(report.qty_books_prlib or 0 for report in book_reports) + \
                       sum(report.qty_books_litres or 0 for report in book_reports) + \
                       sum(report.qty_books_consultant or 0 for report in book_reports) + \
                       sum(report.qty_books_local_library or 0 for report in book_reports)
        website_visits = sum(report.qty_visited_online or 0 for report in visit_reports) + \
                         sum(event.online or 0 for event in events)

        # Заполняем данные для филиала
        ws[f'A{current_row}'] = branch.full_name
        ws[f'B{current_row}'] = children
        ws[f'C{current_row}'] = adults
        ws[f'D{current_row}'] = total_regs
        ws[f'E{current_row}'] = events_count
        ws[f'F{current_row}'] = events_participants
        ws[f'G{current_row}'] = total_visits
        ws[f'H{current_row}'] = books_issued
        ws[f'K{current_row}'] = website_visits

        # Суммируем для итогов
        totals['children'] += children
        totals['adults'] += adults
        totals['total_regs'] += total_regs
        totals['events_count'] += events_count
        totals['events_participants'] += events_participants
        totals['total_visits'] += total_visits
        totals['books_issued'] += books_issued
        totals['website_visits'] += website_visits

        current_row += 1

    # Заполняем строку с итогами (последняя строка в шаблоне)
    ws[f'A{current_row}'] = "ИТОГО:"
    ws[f'B{current_row}'] = totals['children']
    ws[f'C{current_row}'] = totals['adults']
    ws[f'D{current_row}'] = totals['total_regs']
    ws[f'E{current_row}'] = totals['events_count']
    ws[f'F{current_row}'] = totals['events_participants']
    ws[f'G{current_row}'] = totals['total_visits']
    ws[f'H{current_row}'] = totals['books_issued']
    ws[f'K{current_row}'] = totals['website_visits']

    # Сохранение файла
    filename = f"нац_проект_все_филиалы_{year}_{month}.xlsx"
    safe_filename = quote(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
    wb.save(response)
    return response
